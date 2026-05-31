from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession
from io import BytesIO
from datetime import datetime
from fastapi.responses import StreamingResponse

from app.crud import order as crud_order
from app.db.session import get_db
from app.models.order import OrderStatus
from app.schemas.order import OrderCreate, OrderUpdate, OrderOut, OrderListItem
from app.api.deps import get_current_user, require_role
from app.models.user import User, UserRole

router = APIRouter(prefix="/orders", tags=["orders"])


@router.get("", response_model=list[OrderListItem])
async def list_orders(
    skip: int = 0,
    limit: int = 100,
    status_filter: OrderStatus | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return await crud_order.get_all(db, skip=skip, limit=limit, status=status_filter)


@router.get("/stats")
async def orders_stats(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),):
    return await crud_order.get_stats(db)


@router.post("", response_model=OrderOut, status_code=status.HTTP_201_CREATED)
async def create_order(payload: OrderCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),):
    return await crud_order.create(db, payload)


@router.get("/{order_id}", response_model=OrderOut)
async def get_order(order_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),):
    obj = await crud_order.get(db, order_id)
    if not obj:
        raise HTTPException(404, "Order not found")
    return obj


@router.patch("/{order_id}", response_model=OrderOut)
async def update_order(
    order_id: int, payload: OrderUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),
):
    obj = await crud_order.get(db, order_id)
    if not obj:
        raise HTTPException(404, "Order not found")
    return await crud_order.update(db, obj, payload)


@router.delete("/{order_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_order(order_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_role(UserRole.ADMIN)),):
    obj = await crud_order.get(db, order_id)
    if not obj:
        raise HTTPException(404, "Order not found")
    await crud_order.delete(db, obj)


from app.crud import calculation as crud_calc
from app.schemas.calculation import CalculationOut


@router.get("/{order_id}/calculations", response_model=list[CalculationOut])
async def get_order_calculations(order_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user),):
    """История расчетов по заявке."""
    obj = await crud_order.get(db, order_id)
    if not obj:
        raise HTTPException(404, "Order not found")
    return await crud_calc.get_by_order(db, order_id)

@router.get("/my/assigned", response_model=list[OrderListItem])
async def my_assigned_orders(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.DRIVER)),
):
    """Рейсы, назначенные на ТС текущего водителя."""
    return await crud_order.get_for_driver(db, current_user.id)


@router.patch("/{order_id}/status", response_model=OrderOut)
async def driver_change_status(
    order_id: int,
    new_status: OrderStatus,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Смена статуса заявки. Водитель может менять только статусы своих рейсов
    и только в ограниченном наборе (assigned → in_transit → delivered).
    """
    obj = await crud_order.get(db, order_id)
    if not obj:
        raise HTTPException(404, "Order not found")

    # Если это водитель — проверяем, что рейс действительно его
    if current_user.role == UserRole.DRIVER:
        # Находим ТС водителя
        from sqlalchemy import select
        from app.models.vehicle import Vehicle
        v_result = await db.execute(
            select(Vehicle.id).where(Vehicle.driver_id == current_user.id)
        )
        vehicle_id = v_result.scalar_one_or_none()

        if obj.vehicle_id != vehicle_id:
            raise HTTPException(403, "Это не ваш рейс")

        # Разрешенные переходы для водителя
        allowed = {
            OrderStatus.ASSIGNED: [OrderStatus.IN_TRANSIT],
            OrderStatus.IN_TRANSIT: [OrderStatus.DELIVERED],
        }
        if new_status not in allowed.get(obj.status, []):
            raise HTTPException(
                400, f"Нельзя сменить статус с {obj.status} на {new_status}"
            )

    obj.status = new_status
    await db.commit()
    await db.refresh(obj)
    return obj

@router.get("/export/xlsx")
async def export_orders_xlsx(
    status_filter: OrderStatus | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Экспорт списка заявок в Excel-файл."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    orders = await crud_order.get_all(db, limit=10000, status=status_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Заголовок
    headers = [
        "№ заявки", "Откуда", "Куда", "Мили", "Ставка $/mi",
        "Выручка $", "Груз", "Вес (lbs)", "Статус", "Создана",
    ]
    ws.append(headers)

    # Стилизация шапки
    header_fill = PatternFill("solid", fgColor="3B82F6")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные
    status_labels = {
        OrderStatus.DRAFT: "Черновик",
        OrderStatus.ASSIGNED: "Назначен",
        OrderStatus.IN_TRANSIT: "В пути",
        OrderStatus.DELIVERED: "Доставлено",
        OrderStatus.CANCELLED: "Отменен",
    }

    for o in orders:
        revenue = float(o.distance_miles) * float(o.rate_per_mile)
        ws.append([
            o.order_number,
            o.origin_address,
            o.destination_address,
            float(o.distance_miles),
            float(o.rate_per_mile),
            round(revenue, 2),
            o.cargo_type or "",
            float(o.weight_lbs) if o.weight_lbs else "",
            status_labels.get(o.status, o.status),
            o.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    # Авто-ширина колонок (примерная)
    widths = [12, 28, 28, 10, 12, 14, 16, 12, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Закрепляем шапку
    ws.freeze_panes = "A2"

    # Сериализуем в память
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"orders-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/analytics/summary")
async def orders_analytics(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Агрегированная аналитика для дашборда."""
    from sqlalchemy import func, select
    from app.models.order import Order

    # 1. Выручка и кол-во заявок по месяцам
    monthly_q = (
        select(
            func.date_trunc('month', Order.created_at).label('month'),
            func.coalesce(func.sum(Order.distance_miles * Order.rate_per_mile), 0).label('revenue'),
            func.count(Order.id).label('count'),
        )
        .group_by('month')
        .order_by('month')
    )
    monthly_result = await db.execute(monthly_q)
    monthly = [
        {
            "month": row.month.strftime("%Y-%m"),
            "revenue": float(row.revenue),
            "count": row.count,
        }
        for row in monthly_result
    ]

    # 2. Распределение по статусам
    status_q = select(Order.status, func.count(Order.id)).group_by(Order.status)
    status_result = await db.execute(status_q)
    status_counts = {status.value: count for status, count in status_result}

    # 3. Топ-5 маршрутов
    routes_q = (
        select(
            Order.origin_address,
            Order.destination_address,
            func.count(Order.id).label('count'),
            func.coalesce(func.sum(Order.distance_miles * Order.rate_per_mile), 0).label('revenue'),
        )
        .group_by(Order.origin_address, Order.destination_address)
        .order_by(func.count(Order.id).desc())
        .limit(5)
    )
    routes_result = await db.execute(routes_q)
    top_routes = [
        {
            "route": f"{r.origin_address} → {r.destination_address}",
            "count": r.count,
            "revenue": float(r.revenue),
        }
        for r in routes_result
    ]

    return {
        "monthly": monthly,
        "status_counts": status_counts,
        "top_routes": top_routes,
    }